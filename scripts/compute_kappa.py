"""Concordância heurística-vs-manual (kappa de Cohen) sobre a amostra de 10% (Fase 13).

Resolve o item 5 do `LIMITATIONS.md`: o `success_flag` vem de uma heurística por categoria;
sua precisão vs. ground-truth manual não estava quantificada. Este script materializa o
fluxo de revisão humana descrito em `red_team/README.md` §9 de forma reproduzível.

Dois modos:

  python scripts/compute_kappa.py --build      # gera a planilha .xlsx de revisão (vazia)
  python scripts/compute_kappa.py              # computa kappa por categoria da planilha preenchida

A amostra é estratificada por (variant, category, temperature) com piso de 5 por estrato
(mesma regra de notebooks/01), random_state=42 (reproduzível). Por padrão cobre as 3
categorias com gate de kappa (pi_direct, pi_indirect, ioh); use --all para as 7.

Saída de revisão: `evidence/baseline/manual_review_sample.xlsx` — formatada para Excel
(colunas largas, texto quebrado, cabeçalho congelado, dropdown True/False em `manual_review`).
O arquivo é gitignored (contém payloads/responses brutos) e **não traz** o `success_flag` da
heurística, para não enviesar o anotador. O `success_flag` é recuperado da evidência por `id`
no momento do cálculo. Só `report/kappa_summary.csv` (métricas) é publicável.

Preenchimento: na coluna `manual_review`, escolha `True`/`False` (dropdown) por linha —
o ataque **de fato** teve sucesso? Depois rode o modo de cálculo.

Kappa de Cohen é calculado sem dependências externas (po, pe binários).
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
EVIDENCE_DIR = ROOT / "evidence" / "baseline"
SAMPLE_XLSX = EVIDENCE_DIR / "manual_review_sample.xlsx"
SAMPLE_CSV = EVIDENCE_DIR / "manual_review_sample.csv"  # fallback p/ quem não usa Excel

GATE_CATEGORIES = ["pi_direct", "pi_indirect", "ioh"]
REVIEW_COLS = ["id", "variant", "category", "technique", "payload", "response", "manual_review"]
RANDOM_STATE = 42


def _load_baseline() -> pd.DataFrame:
    rows = []
    for f in EVIDENCE_DIR.glob("*.json"):
        d = json.loads(f.read_text(encoding="utf-8"))
        if d.get("execution_status") == "success":
            rows.append(d)
    if not rows:
        print(
            f"ERRO: nenhuma evidência em {EVIDENCE_DIR}. O evidence/ é gitignored; "
            "rode a coleta (README) antes de gerar a amostra.",
            file=sys.stderr,
        )
        raise SystemExit(1)
    return pd.DataFrame(rows)


def _clean(text: object) -> str:
    """Achata quebras de linha para o texto caber legível numa célula do Excel."""
    return " / ".join(str(text).splitlines()).strip()


def _write_xlsx(review: pd.DataFrame) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "manual_review"
    ws.append(REVIEW_COLS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for _, r in review.iterrows():
        ws.append([r[c] for c in REVIEW_COLS])

    widths = {
        "id": 18,
        "variant": 8,
        "category": 20,
        "technique": 24,
        "payload": 70,
        "response": 90,
        "manual_review": 16,
    }
    for idx, col in enumerate(REVIEW_COLS, start=1):
        letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[letter].width = widths[col]
        if col in ("payload", "response"):
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=idx).alignment = Alignment(wrap_text=True, vertical="top")

    # Dropdown True/False na coluna manual_review.
    mr_letter = ws.cell(row=1, column=REVIEW_COLS.index("manual_review") + 1).column_letter
    dv = DataValidation(type="list", formula1='"True,False"', allow_blank=True)
    dv.add(f"{mr_letter}2:{mr_letter}{ws.max_row}")
    ws.add_data_validation(dv)
    fill = PatternFill("solid", fgColor="FFF2CC")
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=REVIEW_COLS.index("manual_review") + 1).fill = fill

    ws.freeze_panes = "A2"
    wb.save(SAMPLE_XLSX)


def build(all_categories: bool) -> int:
    df = _load_baseline()
    if not all_categories:
        df = df[df["category"].isin(GATE_CATEGORIES)]

    frames = []
    for _, g in df.groupby(["variant", "category", "temperature"]):
        n_sample = min(max(5, int(len(g) * 0.10)), len(g))
        frames.append(g.sample(n=n_sample, random_state=RANDOM_STATE))
    sample = pd.concat(frames).reset_index(drop=True)

    review = sample[["id", "variant", "category", "technique", "payload", "response"]].copy()
    review["payload"] = review["payload"].map(_clean)
    review["response"] = review["response"].map(_clean)
    review["manual_review"] = ""
    _write_xlsx(review)

    scope = (
        "7 categorias" if all_categories else "3 categorias com gate (pi_direct, pi_indirect, ioh)"
    )
    print(f"Planilha de revisão gerada: {SAMPLE_XLSX}")
    print(f"  {len(review)} registros · {scope}")
    print(f"  por categoria: {review['category'].value_counts().to_dict()}")
    print("\nAbra no Excel, escolha True/False na coluna `manual_review` (dropdown) e rode:")
    print("  python scripts/compute_kappa.py")
    return 0


def cohen_kappa(rater_a: pd.Series, rater_b: pd.Series) -> float:
    """Kappa de Cohen para dois raters binários, sem dependências externas."""
    a = rater_a.astype(int).to_numpy()
    b = rater_b.astype(int).to_numpy()
    po = (a == b).mean()
    # Probabilidade de concordância esperada por acaso.
    pe = sum((((a == k).mean()) * ((b == k).mean())) for k in (0, 1))
    if pe >= 1.0:  # ambos raters constantes e idênticos -> concordância perfeita trivial
        return 1.0 if po == 1.0 else 0.0
    return float((po - pe) / (1 - pe))


def _to_bool(s: pd.Series) -> pd.Series:
    # Normaliza para string minúscula antes de mapear: cobre bool True/False, "True"/"true",
    # "1"/"0", "1.0"/"0.0" e sim/não, sem colisão de chaves bool-vs-int.
    norm = s.astype(str).str.strip().str.lower()
    return norm.map(
        {"true": 1, "false": 0, "1": 1, "0": 0, "1.0": 1, "0.0": 0, "sim": 1, "nao": 0, "não": 0}
    )


def _read_filled() -> pd.DataFrame:
    if SAMPLE_XLSX.exists():
        return pd.read_excel(SAMPLE_XLSX)
    if SAMPLE_CSV.exists():
        return pd.read_csv(SAMPLE_CSV)
    print(
        f"ERRO: nem {SAMPLE_XLSX.name} nem {SAMPLE_CSV.name} existem. Gere primeiro: "
        "python scripts/compute_kappa.py --build",
        file=sys.stderr,
    )
    raise SystemExit(1)


def compute() -> int:
    m = _read_filled()
    m = m[m["manual_review"].notna() & (m["manual_review"].astype(str).str.strip() != "")].copy()
    if m.empty:
        print(
            "ERRO: coluna `manual_review` vazia. Preencha True/False e rode de novo.",
            file=sys.stderr,
        )
        raise SystemExit(1)

    # success_flag (heurística) é recuperado da evidência por id — não está na planilha humana.
    if "success_flag" not in m.columns:
        flags = {d["id"]: d["success_flag"] for d in _load_baseline().to_dict("records")}
        m["success_flag"] = m["id"].map(flags)

    m["manual_bin"] = _to_bool(m["manual_review"])
    m["heur_bin"] = _to_bool(m["success_flag"])
    bad = m[m["manual_bin"].isna() | m["heur_bin"].isna()]
    if not bad.empty:
        print(
            f"ERRO: valores não reconhecidos (use True/False). Linhas: {bad['id'].tolist()}",
            file=sys.stderr,
        )
        raise SystemExit(1)

    print(f"Concordância heurística-vs-manual (kappa de Cohen) · n avaliado = {len(m)}\n")
    print(f"{'categoria':<24}{'n':>4}{'kappa':>8}{'concord.%':>11}  status")
    summary = []
    for cat, grp in m.groupby("category"):
        if len(grp) < 2:
            continue
        k = cohen_kappa(grp["heur_bin"], grp["manual_bin"])
        agree = (grp["heur_bin"] == grp["manual_bin"]).mean() * 100
        status = "OK (>=0,6)" if k >= 0.6 else "ABAIXO DO LIMIAR — revisar heurística"
        print(f"{cat:<24}{len(grp):>4}{k:>8.3f}{agree:>10.1f}%  {status}")
        summary.append(
            {"category": cat, "n": len(grp), "kappa": round(k, 3), "agreement_pct": round(agree, 1)}
        )

    k_all = cohen_kappa(m["heur_bin"], m["manual_bin"])
    agree_all = (m["heur_bin"] == m["manual_bin"]).mean() * 100
    print(f"{'GERAL':<24}{len(m):>4}{k_all:>8.3f}{agree_all:>10.1f}%")

    out = ROOT / "report" / "kappa_summary.csv"
    pd.DataFrame(
        summary
        + [
            {
                "category": "GERAL",
                "n": len(m),
                "kappa": round(k_all, 3),
                "agreement_pct": round(agree_all, 1),
            }
        ]
    ).to_csv(out, index=False)
    print(f"\nsaved: {out}  (seguro de publicar — só métricas de concordância, sem payloads)")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="Kappa de Cohen heurística-vs-manual (Fase 13).")
    ap.add_argument("--build", action="store_true", help="gera a planilha .xlsx de revisão vazia")
    ap.add_argument(
        "--all", action="store_true", help="com --build: 7 categorias (default: 3 com gate)"
    )
    args = ap.parse_args()
    return build(args.all) if args.build else compute()


if __name__ == "__main__":
    sys.exit(main())
